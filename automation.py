import sys
import time
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
import openpyxl
from selenium.webdriver.common.action_chains import ActionChains
chrome_options = Options()
chrome_options.add_argument("--headless")  # do not show the Chrome window
driver = webdriver.Chrome()
driver.maximize_window()
class Helper:
    @staticmethod
    def get_element(Xpath):
        return driver.find_element(By.XPATH, Xpath)
    @staticmethod
    def goto_rfq_from_pr():
        goto_rfq_from_pr = Helper.get_element('/html/body/div[1]/div/div[2]/div/div/div[3]/div[2]/div/a[2]')
        goto_rfq_from_pr.click()
        time.sleep(5)
    @staticmethod
    def goto_rr():
        goto_rr = Helper.get_element('/html/body/div[1]/div/div[2]/div/div/div[3]/div[2]/div/a[1]')
        goto_rr.click()
        time.sleep(5)
    @staticmethod
    def goto_quote_param():
        goto_quote_param = Helper.get_element('/html/body/div[1]/div/div[2]/div/div/div[3]/div[2]/div/a[3]')
        goto_quote_param.click()
        time.sleep(5)
    @staticmethod
    def goto_quote_comparision():
        goto_quote_comparision = Helper.get_element('/html/body/div[1]/div/div[2]/div/div/div[3]/div[2]/div/a[4]')
        goto_quote_comparision.click()
        time.sleep(5)
    @staticmethod
    def item_tab():
        rr_item = Helper.get_element(
            '/html/body/div[1]/div/div[1]/div[2]/div[4]/div[2]/div/div/div[2]/div[2]/div/div/div[2]/div/div/div['
            '2]/div/div/div[3]/div[2]/div/div/div/div[2]/div/div[1]/div[1]/div[2]/div/a[2]')
        rr_item.click()
        time.sleep(15)
    @staticmethod
    def base_tab():
        rr_base = Helper.get_element(
            '/html/body/div[1]/div/div[1]/div[2]/div[4]/div[2]/div/div/div[2]/div[2]/div/div/div[2]/div/div/div['
            '2]/div/div/div[3]/div[2]/div/div/div/div[2]/div/div[1]/div[1]/div[2]/div/a[1]')
        rr_base.click()
        time.sleep(10)
    @staticmethod
    def order_tab():
        rr_order = Helper.get_element(
            '/html/body/div[1]/div/div[1]/div[2]/div[4]/div[2]/div/div/div[2]/div[2]/div/div/div[2]/div/div/div['
            '2]/div/div/div[3]/div[2]/div/div/div/div[2]/div/div[1]/div[1]/div[2]/div/a[3]')
        rr_order.click()
        time.sleep(10)
    @staticmethod
    def save():
        rr_save = Helper.get_element(
            '/html/body/div[1]/div/div[1]/div[2]/div[4]/div[2]/div/div/div[2]/div[2]/div/div/div[1]/div/div/div['
            '1]/div/div/a[2]')
        rr_save.click()
        time.sleep(10)
class Automation:
    @classmethod
    def get_applicant_name_projectid(cls):
        selected_files = eval(sys.argv[1])

        file_path=selected_files['Applicant name Id']
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        data = {}
        for row in sheet.iter_rows(min_row=5, max_row=11, min_col=2, max_col=4, values_only=True):
            key = str(row[0]).strip() + ' ' + str(row[1]).strip()  # Combine B and C columns as key
            value = str(row[2]).strip()  # D column is the value
            data[key] = value
        data = {key.replace(': None', ''): value for key, value in data.items()}
        data_dict2 = {}
        for row in sheet.iter_rows(min_row=5, max_row=11, min_col=6, max_col=7, values_only=True):
            key = str(row[0]).strip()  # F column is the key
            value = str(row[1]).strip()  # G column is the value
            data_dict2[key] = value
        data_dict2 = {key.replace(':', ''): value for key, value in data_dict2.items()}
        data.update(data_dict2)
        print(data)
        return data
    @staticmethod
    def replenishment_requisition(cost_center_code, supplier_code, applicant_name=None, project_id=None):
        if applicant_name is None:
            applicant_name = Automation.get_applicant_name_projectid()['APPLICANT']
        if project_id is None:
            project_id = Automation.get_applicant_name_projectid()['PROJECT']
        # Step 2
        start_button = Helper.get_element('/html/body/div[1]/div/div[2]/div/div/a')
        start_button.click()
        time.sleep(5)
        search_start = Helper.get_element('//*[@id="search-menu-inputEl"]')
        search_start.send_keys("Replenishment Requisition")
        search_start.send_keys(Keys.RETURN)
        time.sleep(5)
        # Step 3
        rr_combo_box = WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH,
                                            '/html/body/div[1]/div/div[1]/div[2]/div[4]/div[2]/div/div/div[2]/div['
                                            '2]/div/div/div[2]/div/div/div[2]/div/div/div[2]/div/div/div['
                                            '1]/div/div/div[3]/div[1]/div/div[2]')))
        rr_combo_box.click()
        rr_select_project = WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH,
                                            '/html/body/div[9]/div[1]/ul/li[2]')))
        rr_select_project.click()
        # Step 4
        time.sleep(10)
        WebDriverWait(driver, 60).until(EC.presence_of_all_elements_located((By.XPATH,
                                                                             '/html/body/div[1]/div/div[1]/div['
                                                                             '2]/div[4]/div[2]/div/div/div[2]/div['
                                                                             '2]/div/div/div[2]/div/div/div['
                                                                             '2]/div/div/div[2]/div/div/div[2]/div['
                                                                             '2]/div/div[2]/div/div[1]/div/div/div['
                                                                             '1]/div[2]/div/div[1]/div/div/input')))
        rr_project_id = Helper.get_element(
            '/html/body/div[1]/div/div[1]/div[2]/div[4]/div[2]/div/div/div[2]/div[2]/div/div/div[2]/div/div/div['
            '2]/div/div/div[2]/div/div/div[2]/div[2]/div/div[2]/div/div[1]/div/div/div[1]/div[2]/div/div['
            '1]/div/div/input')
        rr_project_id.send_keys(project_id)
        rr_project_id.send_keys(Keys.RETURN)
        time.sleep(20)
        # Step 5
        scroll = Helper.get_element(
            '/html/body/div[1]/div/div[1]/div[2]/div[4]/div[2]/div/div/div[2]/div[2]/div/div/div[2]/div/div/div[2]/div/div/div[2]/div/div/div[2]/div[1]/div[3]')
        ActionChains(driver).move_to_element(scroll).click(scroll).perform()
        release_level_edit = Helper.get_element(
            '/html/body/div[1]/div/div[1]/div[2]/div[4]/div[2]/div/div/div[2]/div[2]/div/div/div[2]/div/div/div[2]/div/div/div[2]/div/div/div[2]/div[2]/div/div[2]/div/div[1]/div/div/div[31]/div[2]/div/div[1]/div/div[1]/input')
        release_level_edit.send_keys('Inserted')
        release_level_edit.send_keys(Keys.RETURN)
        time.sleep(10)
        i = 0
        while True:
            if (i == 1):
                break
            # step 6
            try:
                rr_select_line = Helper.get_element(
                    f'/html/body/div[1]/div/div[1]/div[2]/div[4]/div[2]/div/div/div[2]/div[2]/div/div/div[2]/div/div/div[2]/div/div/div[2]/div/div/div[2]/div[1]/div[1]/div[3]/div/div[2]/table[{i}]')
            except Exception:
                # If no more rows are left
                print("Process Finished:")
                break
            rr_select_line.click()
            time.sleep(15)
            # Change name
            applicant_name_edit = Helper.get_element(
                '/html/body/div[1]/div/div[1]/div[2]/div[4]/div[2]/div/div/div[2]/div[2]/div/div/div[2]/div/div/div[2]/div/div/div[3]/div[2]/div/div/div/div[2]/div/div[2]/div/div/div/div/div/div/div[10]/div/div/div/div[1]/div/div[2]/input')
            applicant_name_edit.send_keys(Keys.CONTROL + "a")
            applicant_name_edit.send_keys(Keys.DELETE)
            applicant_name_edit.send_keys(applicant_name)
            time.sleep(10)
            # Step 7
            Helper.item_tab()
            rr_cost_center_edit = Helper.get_element(
                '/html/body/div[1]/div/div[1]/div[2]/div[4]/div[2]/div/div/div[2]/div[2]/div/div/div[2]/div/div/div[2]/div/div/div[3]/div[2]/div/div/div/div[2]/div/div[2]/div[2]/div/div/div/div/div/div[14]/div/div/div/div[1]/div/div[2]/input')
            rr_cost_center_edit.send_keys(Keys.CONTROL + "a")
            rr_cost_center_edit.send_keys(Keys.DELETE)
            rr_cost_center_edit.send_keys(cost_center_code)
            time.sleep(15)
            # step 8
            Helper.base_tab()
            one_level_approved = Helper.get_element(
                '/html/body/div[1]/div/div[1]/div[2]/div[4]/div[2]/div/div/div[2]/div[2]/div/div/div[2]/div/div/div[2]/div/div/div[3]/div[2]/div/div/div/div[2]/div/div[2]/div/div/div/div/div/div/div[12]/div[1]/div[2]/div[1]/input')
            one_level_approved.send_keys(Keys.CONTROL + "a")
            one_level_approved.send_keys(Keys.DELETE)
            one_level_approved.send_keys("I level approval")
            time.sleep(5)
            # save
            Helper.save()
            # Step 9
            rr_select_line = Helper.get_element(
                '/html/body/div[1]/div/div[1]/div[2]/div[4]/div[2]/div/div/div[2]/div[2]/div/div/div[2]/div/div/div[2]/div/div/div[2]/div/div/div[2]/div[1]/div[1]/div[3]/div/div[2]/table[1]')
            rr_select_line.click()
            time.sleep(10)
            completed = Helper.get_element(
                '/html/body/div[1]/div/div[1]/div[2]/div[4]/div[2]/div/div/div[2]/div[2]/div/div/div[2]/div/div/div[2]/div/div/div[3]/div[2]/div/div/div/div[2]/div/div[2]/div/div/div/div/div/div/div[14]/div[1]/div/span/input')
            completed.click()
            time.sleep(5)
            Helper.order_tab()
            rr_supplier = Helper.get_element(
                '/html/body/div[1]/div/div[1]/div[2]/div[4]/div[2]/div/div/div[2]/div[2]/div/div/div[2]/div/div/div[2]/div/div/div[3]/div[2]/div/div/div/div[2]/div/div[2]/div[2]/div/div/div/div/div/div[5]/div/div/div/div[1]/div/div[2]/input')
            # Step 10
            rr_supplier.send_keys(supplier_code)
            time.sleep(10)
            Helper.save()
            i += 1
        scroll = Helper.get_element(
            '/html/body/div[1]/div/div[1]/div[2]/div[4]/div[2]/div/div/div[2]/div[2]/div/div/div[2]/div/div/div[2]/div/div/div[2]/div/div/div[2]/div[1]/div[3]')
        ActionChains(driver).move_to_element(scroll).click(scroll).perform()
        release_level_edit = Helper.get_element(
            '/html/body/div[1]/div/div[1]/div[2]/div[4]/div[2]/div/div/div[2]/div[2]/div/div/div[2]/div/div/div[2]/div/div/div[2]/div/div/div[2]/div[2]/div/div[2]/div/div[1]/div/div/div[31]/div[2]/div/div[1]/div/div[1]/input')

        release_level_edit.send_keys('Completed')

        release_level_edit.send_keys(Keys.RETURN)

        time.sleep(10)

 
    @staticmethod
    def web_automation_jsp(url, user, psw):
        try:

            driver.get(url)  # Navigate to the .jsp application

            # Step 1: find the username and write the username in the input
            WebDriverWait(driver, 60).until(
                EC.presence_of_all_elements_located((By.XPATH, '//*[@id="j_username-inputEl"]')))

            username = Helper.get_element('//*[@id="j_username-inputEl"]')
            username.send_keys(user)
            password = Helper.get_element('//*[@id="j_password-inputEl"]')
            password.send_keys(psw)
            password.send_keys(Keys.RETURN)
            WebDriverWait(driver, 60).until(
                EC.presence_of_all_elements_located((By.XPATH, '/html/body/div[1]/div/div[2]/div/div/a')))

            # Step 2 -10
            Automation.replenishment_requisition(cost_center_code="110000",
                                                 supplier_code="00000001")
          
        except Exception as e:
            print("An error occurred:", e)

        finally:
            # Close the web driver
            driver.quit()

if __name__ == "__main__":
    jsp_url = "http://192.168.0.237:8080/now/MainDesktop.jsp"
    Automation.web_automation_jsp(url=jsp_url, user="emdx", psw="emdx")
